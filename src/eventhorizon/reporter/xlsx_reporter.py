"""XLSX (Excel) file report exporter."""

import logging
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font

from eventhorizon.utils.severity import map_severity

log = logging.getLogger("EventHorizon.XLSXReporter")

HEADERS = ["Category", "Severity", "File", "Line", "Rule", "Message", "Tool", "Recommendation"]


def export_xlsx(
    findings: List[Dict[str, Any]],
    output_path: Path,
) -> Path:
    """Write findings to an XLSX file with bold headers.

    Args:
        findings: List of finding dicts.
        output_path: Full path for the output XLSX file.

    Returns:
        The path to the written file.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Report"

    bold_font = Font(bold=True)
    for col_num, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font

    for row_num, finding in enumerate(findings, 2):
        ws.cell(row=row_num, column=1, value=finding.get("category", "").capitalize())
        ws.cell(row=row_num, column=2, value=map_severity(finding.get("severity", "info")))
        ws.cell(row=row_num, column=3, value=finding.get("file", ""))
        ws.cell(row=row_num, column=4, value=finding.get("line", ""))
        ws.cell(row=row_num, column=5, value=finding.get("rule", ""))
        ws.cell(row=row_num, column=6, value=finding.get("message", ""))
        ws.cell(row=row_num, column=7, value=finding.get("tool", ""))
        ws.cell(row=row_num, column=8, value=finding.get("recommendation", ""))

    # Auto-fit column widths (approximate)
    for col_num, header in enumerate(HEADERS, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = max_len + 2

    wb.save(str(output_path))
    log.info(f"XLSX report written to {output_path} ({len(findings)} rows)")
    return output_path
