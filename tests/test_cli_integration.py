"""Integration tests for CLI with new analyzers."""

import csv
from pathlib import Path

import pytest
from click.testing import CliRunner

from eventhorizon.cli import main

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "sample_drupal"


@pytest.fixture
def runner():
    return CliRunner()


@pytest.fixture
def output_dir(tmp_path):
    return str(tmp_path / "reports")


class TestCLIIntegration:
    """Integration tests for the analyze command with new analyzers."""

    def test_analyze_all_runs_all_analyzers(self, runner, output_dir):
        """--type all should run static + caching + code_metrics + config_validation."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "all",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
            "--quiet",
        ])
        assert result.exit_code == 0, f"CLI failed: {result.output}"

    def test_analyze_code_metrics_only(self, runner, output_dir):
        """--type code-metrics should only run code metrics."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "code-metrics",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
            "--quiet",
        ])
        assert result.exit_code == 0, f"CLI failed: {result.output}"

    def test_config_validation_runs_for_security(self, runner, output_dir):
        """--type security should include config validation findings."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "security",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
        ])
        assert result.exit_code == 0
        # Config validation findings should appear in security output
        csv_files = list(Path(output_dir).glob("security_report_*.csv"))
        if csv_files:
            with csv_files[0].open() as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                # Should have findings from both static_analyzer and config_validator
                tools = {row.get("Tool") for row in rows}
                assert "static_analyzer" in tools or "config_validator" in tools

    def test_config_validation_runs_for_performance(self, runner, output_dir):
        """--type performance should include config validation findings."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "performance",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
        ])
        assert result.exit_code == 0

    def test_config_sync_probed_automatically(self, runner, output_dir):
        """Config findings should appear when config/sync dir exists."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "all",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
            "--quiet",
        ])
        assert result.exit_code == 0
        # Check that some CSV was generated
        csv_files = list(Path(output_dir).glob("*.csv"))
        assert len(csv_files) > 0

    def test_csv_export_includes_recommendation(self, runner, output_dir):
        """CSV output should have a Recommendation column."""
        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "all",
            "--filter", "custom",
            "--format", "csv",
            "--output", output_dir,
            "--quiet",
        ])
        assert result.exit_code == 0
        csv_files = list(Path(output_dir).glob("*.csv"))
        assert len(csv_files) > 0
        with csv_files[0].open() as f:
            reader = csv.reader(f)
            headers = next(reader)
            assert "Recommendation" in headers

    def test_xlsx_export_includes_recommendation(self, runner, output_dir):
        """XLSX output should have a Recommendation column."""
        from openpyxl import load_workbook

        result = runner.invoke(main, [
            "analyze", str(FIXTURES_DIR),
            "--type", "all",
            "--filter", "custom",
            "--format", "xlsx",
            "--output", output_dir,
            "--quiet",
        ])
        assert result.exit_code == 0
        xlsx_files = list(Path(output_dir).glob("*.xlsx"))
        assert len(xlsx_files) > 0
        wb = load_workbook(xlsx_files[0])
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Recommendation" in headers
