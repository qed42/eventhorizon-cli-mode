"""Tests for the reporter modules."""

import csv
import io
from pathlib import Path

import pytest

SAMPLE_FINDINGS = [
    {
        "tool": "custom",
        "file": "modules/custom/example/example.module",
        "line": 10,
        "severity": "error",
        "message": "Test high severity issue",
        "rule": "test_rule_1",
        "category": "security",
    },
    {
        "tool": "custom",
        "file": "modules/custom/example/example.module",
        "line": 20,
        "severity": "warning",
        "message": "Test medium severity issue",
        "rule": "test_rule_2",
        "category": "performance",
    },
    {
        "tool": "caching_analyzer",
        "file": "modules/custom/example/example.module",
        "line": 30,
        "severity": "info",
        "message": "Test low severity issue",
        "rule": "test_rule_3",
        "category": "performance",
    },
]


class TestCSVReporter:

    def test_creates_csv_file(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "test_report.csv"
        result = export_csv(SAMPLE_FINDINGS, output)

        assert result == output
        assert output.exists()

    def test_csv_has_correct_headers(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "test_report.csv"
        export_csv(SAMPLE_FINDINGS, output)

        with output.open() as f:
            reader = csv.reader(f)
            headers = next(reader)
            assert headers == ["Category", "Severity", "File", "Line", "Rule", "Message", "Tool", "Recommendation"]

    def test_csv_has_correct_row_count(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "test_report.csv"
        export_csv(SAMPLE_FINDINGS, output)

        with output.open() as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert len(rows) == 4  # 1 header + 3 data rows

    def test_csv_severity_mapping(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "test_report.csv"
        export_csv(SAMPLE_FINDINGS, output)

        with output.open() as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            rows = list(reader)
            severities = [row[1] for row in rows]
            assert "High" in severities
            assert "Medium" in severities
            assert "Low" in severities

    def test_csv_creates_parent_dirs(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "nested" / "dir" / "report.csv"
        export_csv(SAMPLE_FINDINGS, output)
        assert output.exists()

    def test_csv_empty_findings(self, tmp_path):
        from eventhorizon.reporter.csv_reporter import export_csv

        output = tmp_path / "empty_report.csv"
        export_csv([], output)
        assert output.exists()

        with output.open() as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert len(rows) == 1  # header only


class TestXLSXReporter:

    def test_creates_xlsx_file(self, tmp_path):
        from eventhorizon.reporter.xlsx_reporter import export_xlsx

        output = tmp_path / "test_report.xlsx"
        result = export_xlsx(SAMPLE_FINDINGS, output)

        assert result == output
        assert output.exists()

    def test_xlsx_has_bold_headers(self, tmp_path):
        from openpyxl import load_workbook

        from eventhorizon.reporter.xlsx_reporter import export_xlsx

        output = tmp_path / "test_report.xlsx"
        export_xlsx(SAMPLE_FINDINGS, output)

        wb = load_workbook(str(output))
        ws = wb.active
        for col in range(1, 9):
            assert ws.cell(row=1, column=col).font.bold is True

    def test_xlsx_row_count(self, tmp_path):
        from openpyxl import load_workbook

        from eventhorizon.reporter.xlsx_reporter import export_xlsx

        output = tmp_path / "test_report.xlsx"
        export_xlsx(SAMPLE_FINDINGS, output)

        wb = load_workbook(str(output))
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 data rows


class TestTerminalReporter:

    def test_returns_severity_counts(self):
        from rich.console import Console

        from eventhorizon.reporter.terminal_reporter import print_summary

        console = Console(file=io.StringIO())
        counts = print_summary(SAMPLE_FINDINGS, "Test", console)

        assert counts["High"] == 1
        assert counts["Medium"] == 1
        assert counts["Low"] == 1

    def test_empty_findings(self):
        from rich.console import Console

        from eventhorizon.reporter.terminal_reporter import print_summary

        console = Console(file=io.StringIO())
        counts = print_summary([], "Test", console)

        assert counts["High"] == 0
        assert counts["Medium"] == 0
        assert counts["Low"] == 0
